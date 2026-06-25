from __future__ import annotations

import json
import importlib
import re
from pathlib import Path

from openpyxl import load_workbook


EXAMPLE_DIR = Path("examples/esercizio_excel_base")
REQUIRED_FILES = {
    "modello_vuoto.xlsx",
    "modello_risolto.xlsx",
    "studente_corretto.xlsx",
    "studente_errori.xlsx",
    "profilo_correzione.ccal",
    "tabella_valutazione_esempio.txt",
    "report_docente_esempio.txt",
    "feedback_studente_esempio.txt",
    "README.md",
}


def test_example_pack_contains_required_files() -> None:
    assert EXAMPLE_DIR.is_dir()
    present = {path.name for path in EXAMPLE_DIR.iterdir() if path.is_file()}
    assert REQUIRED_FILES <= present


def test_example_workbooks_are_loadable_with_openpyxl() -> None:
    for filename in [
        "modello_vuoto.xlsx",
        "modello_risolto.xlsx",
        "studente_corretto.xlsx",
        "studente_errori.xlsx",
    ]:
        workbook = load_workbook(EXAMPLE_DIR / filename, data_only=False)
        try:
            assert "Esercizio" in workbook.sheetnames
        finally:
            workbook.close()


def test_example_profile_is_valid_json_profile() -> None:
    profile = json.loads((EXAMPLE_DIR / "profilo_correzione.ccal").read_text(encoding="utf-8"))

    assert profile["document_type"] == "correction_profile"
    assert profile["cellcheck_format"] == "ccal"
    assert profile["worksheets"]
    assert profile["worksheets"][0]["rules"]


def test_example_text_files_do_not_contain_local_absolute_paths() -> None:
    forbidden = re.compile(r"([A-Za-z]:\\|/Users/|/home/|\\\\)")
    for path in EXAMPLE_DIR.iterdir():
        if path.suffix.lower() in {".txt", ".md", ".ccal"}:
            text = path.read_text(encoding="utf-8")
            assert forbidden.search(text) is None, path.name


def test_example_student_feedback_does_not_expose_reserved_solution_fields() -> None:
    text = (EXAMPLE_DIR / "feedback_studente_esempio.txt").read_text(encoding="utf-8")
    forbidden_fragments = [
        "Formula modello",
        "Valore atteso",
        "expected_formula",
        "expected_value",
        "=C5+D5",
        "=C6*D6",
        "Bilancio",
        "Confermato",
        "Presente",
    ]

    for fragment in forbidden_fragments:
        assert fragment not in text


def test_example_assessment_table_does_not_expose_solutions() -> None:
    text = (EXAMPLE_DIR / "tabella_valutazione_esempio.txt").read_text(encoding="utf-8")
    forbidden_fragments = [
        "Formula modello",
        "Valore atteso",
        "expected_formula",
        "expected_value",
        "=C5+D5",
        "=C6*D6",
        "Bilancio",
        "Confermato",
        "Presente",
    ]

    for fragment in forbidden_fragments:
        assert fragment not in text


def test_example_profile_cells_exist_in_solved_workbook() -> None:
    profile = json.loads((EXAMPLE_DIR / "profilo_correzione.ccal").read_text(encoding="utf-8"))
    workbook = load_workbook(EXAMPLE_DIR / "modello_risolto.xlsx", data_only=False)
    try:
        for worksheet in profile["worksheets"]:
            assert worksheet["sheet_name"] in workbook.sheetnames
            sheet = workbook[worksheet["sheet_name"]]
            for rule in worksheet["rules"]:
                if rule.get("cell"):
                    assert sheet[rule["cell"]] is not None
                if rule.get("range_ref"):
                    assert list(sheet[rule["range_ref"]])
    finally:
        workbook.close()


def test_example_readme_is_present_and_not_empty() -> None:
    readme = EXAMPLE_DIR / "README.md"

    assert readme.exists()
    assert len(readme.read_text(encoding="utf-8").strip()) > 200


def test_existing_manual_workbook_generator_is_importable_without_running_generation() -> None:
    module = importlib.import_module("tools.generate_manual_test_workbooks")

    assert callable(module.main)
