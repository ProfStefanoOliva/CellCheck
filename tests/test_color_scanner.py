from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill

from cellcheck.core import ColorScanner, InvalidColorInputError, WorksheetNotFoundError, parse_color_input
from cellcheck.models import WorkbookFormat


TARGET_FILL = PatternFill(fill_type="solid", fgColor="FFD9D9D9")
OTHER_FILL = PatternFill(fill_type="solid", fgColor="FFFF0000")
INDEXED_TARGET_FILL = PatternFill(fill_type="solid", fgColor=Color(indexed=5))
THEME_TARGET_FILL = PatternFill(fill_type="solid", fgColor=Color(theme=4))


def create_color_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "Header"
    sheet["A1"].fill = TARGET_FILL
    sheet["B2"].fill = TARGET_FILL
    sheet["C3"] = "Other"
    sheet["C3"].fill = OTHER_FILL
    sheet["D4"] = "=SUM(1,1)"
    sheet["D4"].fill = TARGET_FILL

    second_sheet = workbook.create_sheet("Sheet2")
    second_sheet["A1"] = "No match"

    workbook.save(path)
    workbook.close()
    return path


def create_multi_color_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "Target RGB"
    sheet["A1"].fill = TARGET_FILL
    sheet["B2"] = "Other RGB"
    sheet["B2"].fill = OTHER_FILL
    sheet["C3"] = "Indexed target"
    sheet["C3"].fill = INDEXED_TARGET_FILL
    sheet["D4"] = "Theme target"
    sheet["D4"].fill = THEME_TARGET_FILL
    sheet["E5"] = "No fill"

    workbook.save(path)
    workbook.close()
    return path


def test_parse_color_input_accepts_hash_rgb() -> None:
    target = parse_color_input("#D9D9D9")
    assert target.normalized_rgb == "D9D9D9"


def test_parse_color_input_accepts_plain_rgb() -> None:
    target = parse_color_input("D9D9D9")
    assert target.normalized_argb == "FFD9D9D9"


def test_parse_color_input_accepts_argb() -> None:
    target = parse_color_input("FFD9D9D9")
    assert target.normalized_rgb == "D9D9D9"


def test_parse_color_input_accepts_argb_with_ff_alpha() -> None:
    target = parse_color_input("FFFFFF00")
    assert target.normalized_rgb == "FFFF00"
    assert target.normalized_argb == "FFFFFF00"


def test_parse_color_input_normalizes_lowercase() -> None:
    target = parse_color_input("ffd9d9d9")
    assert target.normalized_argb == "FFD9D9D9"


def test_parse_color_input_rejects_empty_string() -> None:
    with pytest.raises(InvalidColorInputError):
        parse_color_input("   ")


def test_parse_color_input_rejects_non_hex_characters() -> None:
    with pytest.raises(InvalidColorInputError):
        parse_color_input("GGD9D9")


def test_parse_color_input_rejects_invalid_lengths() -> None:
    with pytest.raises(InvalidColorInputError):
        parse_color_input("D9D9")


def test_scan_fill_color_finds_colored_cell_in_xlsx(tmp_path: Path) -> None:
    path = create_color_workbook(tmp_path / "sample.xlsx")
    with ColorScanner(path) as scanner:
        result = scanner.scan_fill_color("#D9D9D9")
    assert any(match.cell == "A1" for match in result.matches)


def test_scan_fill_color_finds_empty_but_colored_cell_in_xlsx(tmp_path: Path) -> None:
    path = create_color_workbook(tmp_path / "sample.xlsx")
    with ColorScanner(path) as scanner:
        result = scanner.scan_fill_color("D9D9D9")
    assert any(match.cell == "B2" and match.value_preview is None for match in result.matches)


def test_scan_fill_color_excludes_different_colors(tmp_path: Path) -> None:
    path = create_color_workbook(tmp_path / "sample.xlsx")
    with ColorScanner(path) as scanner:
        result = scanner.scan_fill_color("D9D9D9")
    assert all(match.cell != "C3" for match in result.matches)


def test_scan_fill_color_is_case_insensitive_for_target_comparison(tmp_path: Path) -> None:
    path = create_color_workbook(tmp_path / "sample.xlsx")
    with ColorScanner(path) as scanner:
        result = scanner.scan_fill_color("#d9d9d9")
    assert any(match.cell == "A1" for match in result.matches)


def test_scan_fill_color_returns_correct_sheet_and_cell(tmp_path: Path) -> None:
    path = create_color_workbook(tmp_path / "sample.xlsx")
    with ColorScanner(path) as scanner:
        result = scanner.scan_fill_color("FFD9D9D9")
    first_match = next(match for match in result.matches if match.cell == "A1")
    assert first_match.sheet_name == "Sheet1"


def test_scan_fill_color_supports_single_sheet_filter(tmp_path: Path) -> None:
    path = create_color_workbook(tmp_path / "sample.xlsx")
    with ColorScanner(path) as scanner:
        result = scanner.scan_fill_color("D9D9D9", sheet_names=["Sheet2"])
    assert result.summary.scanned_sheets == ["Sheet2"]
    assert result.summary.matched_cells_count == 0


def test_scan_fill_color_raises_when_sheet_is_missing(tmp_path: Path) -> None:
    path = create_color_workbook(tmp_path / "sample.xlsx")
    with ColorScanner(path) as scanner:
        with pytest.raises(WorksheetNotFoundError):
            scanner.scan_fill_color("D9D9D9", sheet_names=["MissingSheet"])


def test_scan_fill_color_supports_xlsm_without_executing_macros(tmp_path: Path) -> None:
    path = create_color_workbook(tmp_path / "sample.xlsm")
    with ColorScanner(path) as scanner:
        result = scanner.scan_fill_color("D9D9D9")
        assert scanner.workbook_format == WorkbookFormat.XLSM
        assert scanner.macro_enabled is True
    assert result.summary.matched_cells_count >= 1


def test_scan_fill_color_excludes_cells_without_background_fill(tmp_path: Path) -> None:
    path = create_multi_color_workbook(tmp_path / "no-fill.xlsx")
    with ColorScanner(path) as scanner:
        result = scanner.scan_fill_color("D9D9D9")
    assert all(match.cell != "E5" for match in result.matches)


def test_scan_fill_color_handles_indexed_target_color(tmp_path: Path) -> None:
    path = create_multi_color_workbook(tmp_path / "indexed.xlsx")
    with ColorScanner(path) as scanner:
        result = scanner.scan_fill_color("#FFFF00")
    assert any(match.cell == "C3" for match in result.matches)
    assert all(match.cell != "B2" for match in result.matches)


def test_scan_fill_color_handles_theme_color_when_workbook_theme_is_available(tmp_path: Path) -> None:
    path = create_multi_color_workbook(tmp_path / "theme.xlsx")
    with ColorScanner(path) as scanner:
        result = scanner.scan_fill_color("4F81BD")
    assert any(match.cell == "D4" for match in result.matches)


def test_scan_fill_color_with_multiple_colors_only_returns_requested_target(tmp_path: Path) -> None:
    path = create_multi_color_workbook(tmp_path / "mixed.xlsx")
    with ColorScanner(path) as scanner:
        result = scanner.scan_fill_color("D9D9D9")
    matched_cells = {match.cell for match in result.matches}
    assert matched_cells == {"A1"}


def test_scan_fill_color_closes_without_resource_warnings(tmp_path: Path) -> None:
    path = create_color_workbook(tmp_path / "sample.xlsm")
    scanner = ColorScanner(path)
    try:
        result = scanner.scan_fill_color("D9D9D9")
        assert result.summary.workbook_path.endswith("sample.xlsm")
    finally:
        scanner.close()
        scanner.close()
    assert scanner._workbook is None


def test_scan_fill_color_returns_zero_matches_when_no_cells_match(tmp_path: Path) -> None:
    path = create_color_workbook(tmp_path / "sample.xlsx")
    with ColorScanner(path) as scanner:
        result = scanner.scan_fill_color("00FF00")
    assert result.summary.matched_cells_count == 0
    assert result.matches == []
