import os
from pathlib import Path

from types import SimpleNamespace

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import pytest
from openpyxl import Workbook, load_workbook
from PySide6.QtWidgets import QApplication

from cellcheck.ui.i18n import TRANSLATIONS, available_languages, current_language, set_current_language
from cellcheck.ui.workbook_preview import (
    WorkbookPreviewDataSource,
    WorkbookPreviewWindow,
    preview_selection_style_role,
    preview_visual_role,
    workbook_basename,
)
from cellcheck.ui.workbook_preview_navigation import (
    first_cell_from_reference,
    parse_preview_reference,
    resolve_target_sheet_name,
)
from cellcheck.ui.workbook_preview_rule_creation import (
    PreviewRuleMatch,
    PreviewSelectionBounds,
    build_rule_draft_from_preview_cell,
    excel_reference_from_selection,
    find_rules_for_preview_reference,
    suggest_rule_type_from_cell,
)
from cellcheck.ui.workbook_preview_highlights import (
    build_highlighted_cells_map,
    excel_column_label,
    expand_excel_reference,
)
from cellcheck.models import RuleType


def _qapp() -> QApplication:
    return QApplication.instance() or QApplication([])


def create_preview_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Input"
    sheet["A1"] = "Name"
    sheet["B2"] = 42
    sheet["C3"] = "=SUM(B2,8)"
    workbook.create_sheet("Summary")
    workbook.save(path)
    workbook.close()
    return path


def test_preview_window_import_is_available() -> None:
    assert WorkbookPreviewWindow is not None


def test_create_rule_button_available_with_models_context_and_selected_cell(tmp_path: Path) -> None:
    _qapp()
    previous_language = current_language()
    set_current_language("it", persist=False)
    path = create_preview_workbook(tmp_path / "preview.xlsx")
    window = WorkbookPreviewWindow(
        path,
        on_rule_create_requested=lambda _draft: True,
        can_create_rule=True,
    )
    try:
        assert window.table.currentRow() == 0
        assert window.table.currentColumn() == 0
        assert window.create_rule_button.isEnabled() is True
        assert window.create_rule_button.text() == TRANSLATIONS["it"]["workbook_preview.create_rule_from_cell"]
    finally:
        window.close()
        set_current_language(previous_language, persist=False)


def test_remove_rule_button_enabled_only_with_associated_rule(tmp_path: Path) -> None:
    _qapp()
    path = create_preview_workbook(tmp_path / "preview.xlsx")

    def lookup(_sheet_name: str, reference: str) -> list[PreviewRuleMatch]:
        if reference == "A1":
            return [
                PreviewRuleMatch(
                    worksheet_index=0,
                    rule_index=0,
                    rule_id="rule-1",
                    sheet_name="Input",
                    target_ref="A1",
                    match_kind="exact_cell",
                )
            ]
        return []

    window = WorkbookPreviewWindow(
        path,
        on_rule_lookup_requested=lookup,
        on_rule_remove_requested=lambda _sheet_name, _reference: True,
    )
    try:
        window.table.setCurrentCell(0, 0)
        assert window.remove_rule_button.isEnabled() is True

        window.table.setCurrentCell(1, 1)
        assert window.remove_rule_button.isEnabled() is False
    finally:
        window.close()


def test_excel_column_label_converts_expected_indexes() -> None:
    assert excel_column_label(1) == "A"
    assert excel_column_label(26) == "Z"
    assert excel_column_label(27) == "AA"


def test_excel_reference_from_selection_converts_single_cell() -> None:
    bounds = PreviewSelectionBounds(top_row=7, left_column=2, bottom_row=7, right_column=2)

    assert excel_reference_from_selection(bounds) == "B7"


def test_excel_reference_from_selection_converts_contiguous_range() -> None:
    bounds = PreviewSelectionBounds(top_row=1, left_column=1, bottom_row=3, right_column=2)

    assert excel_reference_from_selection(bounds) == "A1:B3"


def test_excel_reference_from_selection_handles_columns_beyond_z() -> None:
    bounds = PreviewSelectionBounds(top_row=12, left_column=27, bottom_row=14, right_column=28)

    assert excel_reference_from_selection(bounds) == "AA12:AB14"


def test_expand_excel_reference_parses_single_cell() -> None:
    assert expand_excel_reference("A1") == {"A1"}


def test_expand_excel_reference_parses_column_beyond_z() -> None:
    assert expand_excel_reference("AA1") == {"AA1"}


def test_expand_excel_reference_parses_range() -> None:
    assert expand_excel_reference("A1:B2") == {"A1", "A2", "B1", "B2"}


def test_parse_preview_reference_parses_single_cell_b7() -> None:
    target = parse_preview_reference("B7")
    assert target.reference == "B7"
    assert target.first_cell == "B7"
    assert target.first_row == 7
    assert target.first_column == 2
    assert target.is_range is False


def test_parse_preview_reference_parses_column_beyond_z() -> None:
    target = parse_preview_reference("AA12")
    assert target.reference == "AA12"
    assert target.first_cell == "AA12"
    assert target.first_row == 12
    assert target.first_column == 27


def test_parse_preview_reference_parses_range_h7_k14() -> None:
    target = parse_preview_reference("H7:K14")
    assert target.reference == "H7:K14"
    assert target.first_cell == "H7"
    assert target.first_row == 7
    assert target.first_column == 8
    assert target.is_range is True


def test_first_cell_from_reference_returns_range_anchor() -> None:
    assert first_cell_from_reference("H7:K14") == "H7"


def test_parse_preview_reference_rejects_invalid_cell() -> None:
    with pytest.raises(Exception):
        parse_preview_reference("INVALID!")


def test_parse_preview_reference_rejects_invalid_range() -> None:
    with pytest.raises(Exception):
        parse_preview_reference("A1::B2")


def test_suggest_rule_type_from_formula_cell() -> None:
    assert (
        suggest_rule_type_from_cell(has_formula=True, formula_text="=SUM(A1:A3)", value=None)
        == RuleType.FORMULA_EXACT
    )


def test_suggest_rule_type_from_number_cell() -> None:
    assert (
        suggest_rule_type_from_cell(has_formula=False, formula_text=None, value=42)
        == RuleType.NUMERIC_VALUE
    )


def test_suggest_rule_type_from_text_cell() -> None:
    assert (
        suggest_rule_type_from_cell(has_formula=False, formula_text=None, value="Hello")
        == RuleType.TEXT_VALUE
    )


def test_suggest_rule_type_avoids_fragile_boolean_guess() -> None:
    assert suggest_rule_type_from_cell(has_formula=False, formula_text=None, value=True) is None


def test_preview_visual_role_distinguishes_selected_profile_cell() -> None:
    assert (
        preview_visual_role(is_highlighted=True, is_report_target=False, is_selected=True)
        == "selected_profile_highlight"
    )
    assert (
        preview_visual_role(is_highlighted=True, is_report_target=False, is_selected=False)
        == "profile_highlight"
    )


def test_preview_selection_style_role_separates_profile_from_selected_profile() -> None:
    assert (
        preview_selection_style_role("profile_highlight", is_selected=False)
        == "profile_highlight"
    )
    assert (
        preview_selection_style_role("profile_highlight", is_selected=True)
        == "selected_profile_highlight"
    )
    assert preview_selection_style_role("profile_highlight", is_selected=True) != "profile_highlight"


def test_preview_selection_style_role_uses_manual_selection_family_for_normal_cells() -> None:
    assert preview_selection_style_role("default", is_selected=True) == "manual_selection"


def test_preview_visual_role_keeps_report_target_priority() -> None:
    assert (
        preview_visual_role(is_highlighted=True, is_report_target=True, is_selected=True)
        == "report_target"
    )


def test_preview_selection_style_role_keeps_report_target_priority() -> None:
    assert (
        preview_selection_style_role("report_target", is_selected=True)
        == "report_target_selection"
    )


def test_build_rule_draft_from_selected_cell() -> None:
    draft = build_rule_draft_from_preview_cell(
        sheet_name="Input",
        reference="b2",
        has_formula=False,
        formula_text=None,
        value=42,
        required_activity="Complete the indicated cell.",
    )

    assert draft.sheet_name == "Input"
    assert draft.cell == "B2"
    assert draft.range_ref is None
    assert draft.suggested_rule_type == RuleType.NUMERIC_VALUE
    assert draft.expected_value == 42
    assert draft.expected_formula is None
    assert draft.required_activity == "Complete the indicated cell."


def test_build_rule_draft_from_selected_range() -> None:
    draft = build_rule_draft_from_preview_cell(
        sheet_name="Input",
        reference="A1:B3",
        has_formula=True,
        formula_text="=SUM(A1:A3)",
        value=None,
    )

    assert draft.sheet_name == "Input"
    assert draft.cell is None
    assert draft.range_ref == "A1:B3"
    assert draft.suggested_rule_type == RuleType.MANUAL_REVIEW
    assert draft.expected_formula is None
    assert draft.expected_value is None


def test_find_rules_for_preview_reference_matches_single_cell() -> None:
    profile = SimpleNamespace(
        worksheets=[
            SimpleNamespace(
                sheet_name="Input",
                rules=[SimpleNamespace(id="rule-1", cell="AA12", range_ref=None)],
            )
        ]
    )

    matches = find_rules_for_preview_reference(profile, "Input", "AA12")

    assert len(matches) == 1
    assert matches[0].rule_id == "rule-1"
    assert matches[0].target_ref == "AA12"
    assert matches[0].match_kind == "exact_cell"


def test_find_rules_for_preview_reference_matches_exact_range() -> None:
    profile = SimpleNamespace(
        worksheets=[
            SimpleNamespace(
                sheet_name="Input",
                rules=[SimpleNamespace(id="rule-2", cell=None, range_ref="A1:B3")],
            )
        ]
    )

    matches = find_rules_for_preview_reference(profile, "Input", "A1:B3")

    assert len(matches) == 1
    assert matches[0].rule_id == "rule-2"
    assert matches[0].match_kind == "exact_range"


def test_find_rules_for_preview_reference_matches_cell_inside_range() -> None:
    profile = SimpleNamespace(
        worksheets=[
            SimpleNamespace(
                sheet_name="Input",
                rules=[SimpleNamespace(id="rule-3", cell=None, range_ref="AA12:AB14")],
            )
        ]
    )

    matches = find_rules_for_preview_reference(profile, "Input", "AB13")

    assert len(matches) == 1
    assert matches[0].rule_id == "rule-3"
    assert matches[0].target_ref == "AA12:AB14"
    assert matches[0].match_kind == "cell_inside_range"


def test_find_rules_for_preview_reference_returns_empty_without_match() -> None:
    profile = SimpleNamespace(
        worksheets=[
            SimpleNamespace(
                sheet_name="Input",
                rules=[SimpleNamespace(id="rule-1", cell="A1", range_ref=None)],
            )
        ]
    )

    assert find_rules_for_preview_reference(profile, "Input", "B2") == []


def test_find_rules_for_preview_reference_returns_multiple_matches() -> None:
    profile = SimpleNamespace(
        worksheets=[
            SimpleNamespace(
                sheet_name="Input",
                rules=[
                    SimpleNamespace(id="rule-1", cell="B2", range_ref=None),
                    SimpleNamespace(id="rule-2", cell=None, range_ref="A1:B3"),
                ],
            )
        ]
    )

    matches = find_rules_for_preview_reference(profile, "Input", "B2")

    assert [match.rule_id for match in matches] == ["rule-1", "rule-2"]


def test_resolve_target_sheet_name_returns_requested_sheet() -> None:
    assert resolve_target_sheet_name(["Input", "Summary"], "Summary", "Input") == "Summary"


def test_resolve_target_sheet_name_returns_none_for_missing_sheet() -> None:
    assert resolve_target_sheet_name(["Input", "Summary"], "Missing", "Input") is None


def test_workbook_basename_extracts_filename_from_windows_path() -> None:
    assert workbook_basename(r"C:\Classi\Compito\Studente_01.xlsx") == "Studente_01.xlsx"


def test_preview_data_source_reads_sheet_names(tmp_path: Path) -> None:
    path = create_preview_workbook(tmp_path / "preview.xlsx")
    source = WorkbookPreviewDataSource(path)
    try:
        assert source.sheet_names == ["Input", "Summary"]
    finally:
        source.close()


def test_build_highlighted_cells_map_collects_cells_from_profile_rules() -> None:
    profile = SimpleNamespace(
        worksheets=[
            SimpleNamespace(
                sheet_name="Input",
                rules=[
                    SimpleNamespace(cell="A1"),
                    SimpleNamespace(cell="B2:C2"),
                ],
            )
        ]
    )

    highlighted = build_highlighted_cells_map(profile)

    assert highlighted == {"Input": {"A1", "B2", "C2"}}


def test_build_highlighted_cells_map_keeps_sheets_separate() -> None:
    profile = SimpleNamespace(
        worksheets=[
            SimpleNamespace(sheet_name="Input", rules=[SimpleNamespace(cell="A1")]),
            SimpleNamespace(sheet_name="Summary", rules=[SimpleNamespace(cell="A1:B1")]),
        ]
    )

    highlighted = build_highlighted_cells_map(profile)

    assert highlighted["Input"] == {"A1"}
    assert highlighted["Summary"] == {"A1", "B1"}


def test_build_highlighted_cells_map_returns_empty_for_missing_profile() -> None:
    assert build_highlighted_cells_map(None) == {}
    assert build_highlighted_cells_map(SimpleNamespace(worksheets=[])) == {}


def test_preview_data_source_reads_value_and_formula(tmp_path: Path) -> None:
    path = create_preview_workbook(tmp_path / "preview.xlsx")
    previous_language = current_language()
    set_current_language("it", persist=False)
    source = WorkbookPreviewDataSource(path)
    try:
        cell = source.get_cell_data("Input", 3, 3)
        assert cell.cell_ref == "C3"
        assert cell.has_formula is True
        assert cell.formula_text == "=SUM(B2,8)"
        assert cell.display_value == "=SUM(B2,8)"
        assert cell.cached_value == ""
        assert cell.value_text == "Valore calcolato non disponibile nel file."
    finally:
        source.close()
        set_current_language(previous_language, persist=False)


def test_preview_data_source_marks_cells_without_formula_cleanly(tmp_path: Path) -> None:
    path = create_preview_workbook(tmp_path / "preview.xlsx")
    previous_language = current_language()
    set_current_language("it", persist=False)
    source = WorkbookPreviewDataSource(path)
    try:
        cell = source.get_cell_data("Input", 2, 2)
        assert cell.cell_ref == "B2"
        assert cell.display_value == "42"
        assert cell.cached_value == "42"
        assert cell.value_text == "42"
        assert cell.formula_text == "Nessuna formula"
    finally:
        source.close()
        set_current_language(previous_language, persist=False)


def test_preview_data_source_marks_highlighted_cells(tmp_path: Path) -> None:
    path = create_preview_workbook(tmp_path / "preview.xlsx")
    source = WorkbookPreviewDataSource(path, highlighted_cells_by_sheet={"Input": {"B2", "C3"}})
    try:
        highlighted_cell = source.get_cell_data("Input", 2, 2)
        normal_cell = source.get_cell_data("Input", 1, 1)
        assert highlighted_cell.is_highlighted is True
        assert highlighted_cell.is_report_target is False
        assert highlighted_cell.visual_role == "profile_highlight"
        assert normal_cell.is_highlighted is False
        assert normal_cell.visual_role == "default"
    finally:
        source.close()


def test_preview_data_source_marks_report_target_cells_distinctly(tmp_path: Path) -> None:
    path = create_preview_workbook(tmp_path / "preview.xlsx")
    source = WorkbookPreviewDataSource(path, highlighted_cells_by_sheet={"Input": {"B2"}})
    try:
        source.set_report_target("Input", "C3", "C3")
        target_cell = source.get_cell_data("Input", 3, 3)
        assert target_cell.is_highlighted is False
        assert target_cell.is_report_target is True
        assert target_cell.visual_role == "report_target"
    finally:
        source.close()


def test_report_target_precedes_profile_highlight_when_both_apply(tmp_path: Path) -> None:
    path = create_preview_workbook(tmp_path / "preview.xlsx")
    source = WorkbookPreviewDataSource(path, highlighted_cells_by_sheet={"Input": {"B2"}})
    try:
        source.set_report_target("Input", "B2", "B2")
        target_cell = source.get_cell_data("Input", 2, 2)
        assert target_cell.is_highlighted is True
        assert target_cell.is_report_target is True
        assert target_cell.visual_role == "report_target"
    finally:
        source.close()


def test_report_target_persists_after_navigation_state_update(tmp_path: Path) -> None:
    path = create_preview_workbook(tmp_path / "preview.xlsx")
    source = WorkbookPreviewDataSource(path)
    try:
        source.set_report_target("Input", "H7:K14", "H7")
        assert source.has_report_target() is True
        assert "H7" in source.report_target_cells_by_sheet["Input"]
    finally:
        source.close()


def test_preview_data_source_keeps_truly_empty_cell_empty(tmp_path: Path) -> None:
    path = create_preview_workbook(tmp_path / "preview.xlsx")
    previous_language = current_language()
    set_current_language("it", persist=False)
    source = WorkbookPreviewDataSource(path)
    try:
        cell = source.get_cell_data("Input", 4, 4)
        assert cell.display_value == ""
        assert cell.cached_value == ""
        assert cell.value_text == ""
        assert cell.formula_text == "Nessuna formula"
    finally:
        source.close()
        set_current_language(previous_language, persist=False)


def test_preview_translation_keys_exist_for_all_languages() -> None:
    keys = [
        "navigator.preview_workbook",
        "workbook_preview.title",
        "workbook_preview.action",
        "workbook_preview.student_action",
        "workbook_preview.create_rule",
        "workbook_preview.create_rule_from_cell",
        "workbook_preview.create_rule_from_range",
        "workbook_preview.create_rule_tooltip",
        "workbook_preview.remove_rule",
        "workbook_preview.remove_rule_from_cell",
        "workbook_preview.remove_rule_from_range",
        "workbook_preview.remove_rule_tooltip",
        "workbook_preview.no_associated_rule_title",
        "workbook_preview.no_associated_rule_message",
        "workbook_preview.remove_rule_confirm",
        "workbook_preview.rule_removed_title",
        "workbook_preview.rule_removed_message",
        "workbook_preview.multiple_rules_title",
        "workbook_preview.multiple_rules_message",
        "workbook_preview.solution_rule_tooltip",
        "workbook_preview.student_rule_tooltip",
        "workbook_preview.no_selection_title",
        "workbook_preview.no_selection_message",
        "workbook_preview.non_contiguous_selection",
        "workbook_preview.profile_unavailable_title",
        "workbook_preview.profile_unavailable_message",
        "workbook_preview.reference_models_required_title",
        "workbook_preview.reference_models_required_message",
        "workbook_preview.rule_created_title",
        "workbook_preview.rule_created_message",
        "workbook_preview.student_rule_warning_title",
        "workbook_preview.student_rule_warning_message",
        "workbook_preview.use_solution_hint",
        "workbook_preview.required_activity_default",
        "workbook_preview.sheet",
        "workbook_preview.cell",
        "workbook_preview.value",
        "workbook_preview.formula",
        "workbook_preview.no_formula",
        "workbook_preview.calculated_value_unavailable",
        "workbook_preview.file_unavailable_title",
        "workbook_preview.file_unavailable_message",
        "workbook_preview.student_unavailable_title",
        "workbook_preview.no_student_selected",
        "workbook_preview.report_student_unavailable",
        "workbook_preview.open_cell_in_preview",
        "workbook_preview.no_report_row_selected",
        "workbook_preview.sheet_not_found_title",
        "workbook_preview.sheet_not_found_message",
        "workbook_preview.invalid_reference_title",
        "workbook_preview.invalid_cell_reference",
        "workbook_preview.invalid_range_title",
        "workbook_preview.invalid_range_reference",
        "workbook_preview.highlighted_cells",
        "workbook_preview.highlighted_cells_legend",
        "workbook_preview.report_target_cell_legend",
        "workbook_preview.report_target_range_legend",
        "workbook_preview.no_profile_cells",
        "workbook_preview.sheet_too_large",
    ]
    for language_code, _label in available_languages():
        for key in keys:
            assert key in TRANSLATIONS[language_code]


def test_preview_data_source_does_not_modify_workbook_file(tmp_path: Path) -> None:
    path = create_preview_workbook(tmp_path / "preview.xlsx")
    original_bytes = path.read_bytes()

    source = WorkbookPreviewDataSource(path)
    try:
        source.load_sheet("Input")
        source.get_cell_data("Input", 1, 1)
    finally:
        source.close()

    reopened = load_workbook(path, read_only=True, data_only=False)
    try:
        assert reopened["Input"]["C3"].value == "=SUM(B2,8)"
    finally:
        reopened.close()
    assert path.read_bytes() == original_bytes
