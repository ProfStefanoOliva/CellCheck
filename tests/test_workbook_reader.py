from pathlib import Path

import pytest
from openpyxl import Workbook

from cellcheck.core import (
    CellReadError,
    UnsupportedWorkbookFormatError,
    WorkbookReader,
    WorksheetNotFoundError,
)
from cellcheck.models import WorkbookFormat


def create_workbook_file(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "Hello"
    sheet["B2"] = 123
    sheet["C3"] = "=SUM(B2,2)"
    workbook.create_sheet("Summary")
    workbook.save(path)
    workbook.close()
    return path


def test_recognizes_xlsx(tmp_path: Path) -> None:
    path = create_workbook_file(tmp_path / "sample.xlsx")
    with WorkbookReader(path) as reader:
        assert reader.workbook_format == WorkbookFormat.XLSX


def test_recognizes_xlsm(tmp_path: Path) -> None:
    path = create_workbook_file(tmp_path / "sample.xlsm")
    with WorkbookReader(path) as reader:
        assert reader.workbook_format == WorkbookFormat.XLSM
        assert reader.macro_enabled is True


def test_rejects_xls_extension(tmp_path: Path) -> None:
    path = tmp_path / "sample.xls"
    path.write_text("not-an-excel-file", encoding="utf-8")
    with pytest.raises(UnsupportedWorkbookFormatError):
        WorkbookReader(path)


def test_rejects_csv_extension(tmp_path: Path) -> None:
    path = tmp_path / "sample.csv"
    path.write_text("a,b,c", encoding="utf-8")
    with pytest.raises(UnsupportedWorkbookFormatError):
        WorkbookReader(path)


def test_get_workbook_info_returns_sheet_names(tmp_path: Path) -> None:
    path = create_workbook_file(tmp_path / "sample.xlsx")
    with WorkbookReader(path) as reader:
        info = reader.get_workbook_info()
        assert info.sheet_names == ["Sheet1", "Summary"]


def test_get_workbook_info_returns_correct_format(tmp_path: Path) -> None:
    path = create_workbook_file(tmp_path / "sample.xlsm")
    with WorkbookReader(path) as reader:
        info = reader.get_workbook_info()
        assert info.workbook_format == WorkbookFormat.XLSM


def test_get_workbook_info_returns_macro_enabled_flag(tmp_path: Path) -> None:
    xlsx_path = create_workbook_file(tmp_path / "sample.xlsx")
    xlsm_path = create_workbook_file(tmp_path / "sample.xlsm")

    with WorkbookReader(xlsx_path) as xlsx_reader:
        assert xlsx_reader.get_workbook_info().macro_enabled is False

    with WorkbookReader(xlsm_path) as xlsm_reader:
        assert xlsm_reader.get_workbook_info().macro_enabled is True


def test_get_worksheet_info_returns_dimensions(tmp_path: Path) -> None:
    path = create_workbook_file(tmp_path / "sample.xlsx")
    with WorkbookReader(path) as reader:
        info = reader.get_worksheet_info("Sheet1")
        assert info.max_row == 3
        assert info.max_column == 3


def test_get_cell_snapshot_reads_simple_value(tmp_path: Path) -> None:
    path = create_workbook_file(tmp_path / "sample.xlsx")
    with WorkbookReader(path) as reader:
        snapshot = reader.get_cell_snapshot("Sheet1", "A1")
        assert snapshot.value == "Hello"
        assert snapshot.has_formula is False


def test_get_cell_snapshot_recognizes_formula_when_data_only_false(tmp_path: Path) -> None:
    path = create_workbook_file(tmp_path / "sample.xlsx")
    with WorkbookReader(path, data_only=False) as reader:
        snapshot = reader.get_cell_snapshot("Sheet1", "C3")
        assert snapshot.has_formula is True
        assert snapshot.formula == "=SUM(B2,2)"


def test_get_cell_snapshot_does_not_recalculate_formulas(tmp_path: Path) -> None:
    path = create_workbook_file(tmp_path / "sample.xlsx")
    with WorkbookReader(path, data_only=True) as reader:
        snapshot = reader.get_cell_snapshot("Sheet1", "C3")
        assert snapshot.has_formula is False
        assert snapshot.formula is None
        assert snapshot.value is None


def test_missing_worksheet_raises_error(tmp_path: Path) -> None:
    path = create_workbook_file(tmp_path / "sample.xlsx")
    with WorkbookReader(path) as reader:
        with pytest.raises(WorksheetNotFoundError):
            reader.get_worksheet_info("MissingSheet")


def test_invalid_cell_reference_raises_error(tmp_path: Path) -> None:
    path = create_workbook_file(tmp_path / "sample.xlsx")
    with WorkbookReader(path) as reader:
        with pytest.raises(CellReadError):
            reader.get_cell_snapshot("Sheet1", "INVALID!")


def test_context_manager_closes_without_errors(tmp_path: Path) -> None:
    path = create_workbook_file(tmp_path / "sample.xlsx")
    reader = WorkbookReader(path)
    with reader as managed_reader:
        assert managed_reader.get_workbook_info().filename == "sample.xlsx"
    assert reader._workbook is None
