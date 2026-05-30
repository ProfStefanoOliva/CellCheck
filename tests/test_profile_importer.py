from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from cellcheck.core import ProfileImporter, WorkbookStructureMismatchError
from cellcheck.models import ProfileImportOptions, RuleType, WorkbookFormat


TARGET_FILL = PatternFill(fill_type="solid", fgColor="FFD9D9D9")
OTHER_FILL = PatternFill(fill_type="solid", fgColor="FFFF0000")


def create_template_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"].fill = TARGET_FILL
    sheet["B2"].fill = TARGET_FILL
    sheet["C3"].fill = TARGET_FILL
    sheet["D4"].fill = TARGET_FILL

    second_sheet = workbook.create_sheet("Sheet2")
    second_sheet["A1"].fill = TARGET_FILL

    workbook.save(path)
    workbook.close()
    return path


def create_solution_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "=SUM(1,2)"
    sheet["B2"] = 42
    sheet["C3"] = "Approved"
    sheet["D4"] = None

    second_sheet = workbook.create_sheet("Sheet2")
    second_sheet["A1"] = "Second sheet"

    workbook.save(path)
    workbook.close()
    return path


def create_multi_color_template_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"].fill = TARGET_FILL
    sheet["B2"].fill = OTHER_FILL
    sheet["C3"].fill = TARGET_FILL
    sheet["D4"].fill = OTHER_FILL

    workbook.save(path)
    workbook.close()
    return path


def build_options(**overrides) -> ProfileImportOptions:
    payload = {
        "exercise_name": "Budget Exercise",
        "max_grade": 30.0,
        "target_color": "D9D9D9",
        "default_weight": 2.5,
        "include_empty_solution_cells": True,
    }
    payload.update(overrides)
    return ProfileImportOptions(**payload)


def test_import_profile_generates_valid_profile_from_xlsx(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsx")

    result = ProfileImporter().import_profile(empty_path, solution_path, build_options())

    assert result.profile.exercise_name == "Budget Exercise"
    assert len(result.profile.worksheets) == 2


def test_import_profile_uses_target_color_selected_by_user(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsx")

    result = ProfileImporter().import_profile(
        empty_path,
        solution_path,
        build_options(target_color="#D9D9D9"),
    )

    assert result.summary.target_rgb == "D9D9D9"
    assert result.summary.target_argb == "FFD9D9D9"


def test_import_profile_extracts_only_cells_of_the_requested_target_color(tmp_path: Path) -> None:
    empty_path = create_multi_color_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsx")

    result = ProfileImporter().import_profile(
        empty_path,
        solution_path,
        build_options(target_color="D9D9D9"),
    )

    sheet1_rules = {rule.cell for worksheet in result.profile.worksheets for rule in worksheet.rules}

    assert "A1" in sheet1_rules
    assert "C3" in sheet1_rules
    assert "B2" not in sheet1_rules
    assert "D4" not in sheet1_rules


def test_import_profile_creates_formula_exact_rule(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsx")

    result = ProfileImporter().import_profile(empty_path, solution_path, build_options())
    sheet1_rules = {rule.cell: rule for rule in result.profile.worksheets[0].rules}

    assert sheet1_rules["A1"].rule_type == RuleType.FORMULA_EXACT
    assert sheet1_rules["A1"].expected_formula == "=SUM(1,2)"


def test_import_profile_creates_numeric_value_rule(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsx")

    result = ProfileImporter().import_profile(empty_path, solution_path, build_options())
    sheet1_rules = {rule.cell: rule for rule in result.profile.worksheets[0].rules}

    assert sheet1_rules["B2"].rule_type == RuleType.NUMERIC_VALUE
    assert sheet1_rules["B2"].expected_value == 42


def test_import_profile_creates_text_value_rule(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsx")

    result = ProfileImporter().import_profile(empty_path, solution_path, build_options())
    sheet1_rules = {rule.cell: rule for rule in result.profile.worksheets[0].rules}

    assert sheet1_rules["C3"].rule_type == RuleType.TEXT_VALUE
    assert sheet1_rules["C3"].expected_value == "Approved"


def test_import_profile_creates_manual_review_for_empty_solution_cell(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsx")

    result = ProfileImporter().import_profile(empty_path, solution_path, build_options())
    sheet1_rules = {rule.cell: rule for rule in result.profile.worksheets[0].rules}

    assert sheet1_rules["D4"].rule_type == RuleType.MANUAL_REVIEW


def test_import_profile_ignores_empty_solution_cell_when_disabled(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsx")

    result = ProfileImporter().import_profile(
        empty_path,
        solution_path,
        build_options(include_empty_solution_cells=False),
    )
    sheet1_rules = {rule.cell: rule for rule in result.profile.worksheets[0].rules}

    assert "D4" not in sheet1_rules


def test_import_profile_assigns_default_weight(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsx")

    result = ProfileImporter().import_profile(empty_path, solution_path, build_options())

    assert all(rule.weight == 2.5 for worksheet in result.profile.worksheets for rule in worksheet.rules)


def test_import_profile_sets_max_grade(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsx")

    result = ProfileImporter().import_profile(
        empty_path,
        solution_path,
        build_options(max_grade=18.0),
    )

    assert result.profile.max_grade == 18.0


def test_import_profile_groups_rules_by_sheet(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsx")

    result = ProfileImporter().import_profile(empty_path, solution_path, build_options())

    sheet_names = [worksheet.sheet_name for worksheet in result.profile.worksheets]
    assert sheet_names == ["Sheet1", "Sheet2"]


def test_import_profile_supports_sheet_filter(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsx")

    result = ProfileImporter().import_profile(
        empty_path,
        solution_path,
        build_options(sheet_names=["Sheet2"]),
    )

    assert [worksheet.sheet_name for worksheet in result.profile.worksheets] == ["Sheet2"]


def test_import_profile_raises_if_sheet_is_missing_in_solution(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    workbook.save(tmp_path / "solution.xlsx")
    workbook.close()

    try:
        ProfileImporter().import_profile(empty_path, tmp_path / "solution.xlsx", build_options())
    except WorkbookStructureMismatchError:
        pass
    else:
        raise AssertionError("missing solution sheet should raise WorkbookStructureMismatchError")


def test_import_profile_supports_xlsm_without_executing_macros(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsm")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsm")

    result = ProfileImporter().import_profile(empty_path, solution_path, build_options())

    assert result.summary.workbook_format == WorkbookFormat.XLSM


def test_import_profile_sets_macro_enabled_when_any_workbook_is_xlsm(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsm")

    result = ProfileImporter().import_profile(empty_path, solution_path, build_options())

    assert result.profile.macro_enabled is True


def test_import_profile_does_not_modify_original_workbooks(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsx")

    ProfileImporter().import_profile(empty_path, solution_path, build_options())

    empty_workbook = load_workbook(empty_path)
    solution_workbook = load_workbook(solution_path, data_only=False)
    try:
        assert empty_workbook["Sheet1"]["A1"].fill.fgColor.rgb == "FFD9D9D9"
        assert solution_workbook["Sheet1"]["A1"].value == "=SUM(1,2)"
    finally:
        empty_workbook.close()
        solution_workbook.close()


def test_import_profile_produces_coherent_summary_counts(tmp_path: Path) -> None:
    empty_path = create_template_workbook(tmp_path / "empty.xlsx")
    solution_path = create_solution_workbook(tmp_path / "solution.xlsx")

    result = ProfileImporter().import_profile(empty_path, solution_path, build_options())

    assert result.summary.generated_rules_count == 5
    assert result.summary.formula_rules_count == 1
    assert result.summary.numeric_rules_count == 1
    assert result.summary.text_rules_count == 2
    assert result.summary.manual_review_rules_count == 1
