from pathlib import Path

from openpyxl import Workbook, load_workbook

from cellcheck.core import CorrectionEngine
from cellcheck.models import (
    CellCorrectionResult,
    CorrectionProfile,
    CorrectionRule,
    ResultStatus,
    RuleType,
    ToleranceConfig,
    ToleranceMode,
    WorkbookFormat,
    WorksheetProfile,
)


def create_student_workbook(path: Path, extension: str = ".xlsx") -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "=SUM(1,2)"
    sheet["B2"] = 10
    sheet["C3"] = "  Hello  "
    sheet["D4"] = "Filled"
    sheet["E5"] = None
    second_sheet = workbook.create_sheet("Sheet2")
    second_sheet["A1"] = "Second"
    workbook.save(path.with_suffix(extension))
    workbook.close()
    return path.with_suffix(extension)


def build_rule(
    *,
    rule_id: str,
    cell: str,
    rule_type: RuleType,
    expected_formula: str | None = None,
    expected_value=None,
    weight: float = 2.0,
    enabled: bool = True,
    tolerance: ToleranceConfig | None = None,
    sheet_name: str = "Sheet1",
    range_ref: str | None = None,
) -> CorrectionRule:
    return CorrectionRule(
        id=rule_id,
        sheet_name=sheet_name,
        cell=None if range_ref else cell,
        range_ref=range_ref,
        rule_type=rule_type,
        expected_formula=expected_formula,
        expected_value=expected_value,
        weight=weight,
        enabled=enabled,
        tolerance=tolerance,
        teacher_note="",
    )


def build_profile(rules: list[CorrectionRule], max_grade: float = 30.0) -> CorrectionProfile:
    worksheets: dict[str, list[CorrectionRule]] = {}
    for rule in rules:
        worksheets.setdefault(rule.sheet_name, []).append(rule)

    return CorrectionProfile(
        minimum_cellcheck_version="0.7.0",
        exercise_name="Budget Exercise",
        max_grade=max_grade,
        worksheets=[
            WorksheetProfile(sheet_name=sheet_name, rules=sheet_rules)
            for sheet_name, sheet_rules in worksheets.items()
        ],
    )


def get_result(report, rule_id: str) -> CellCorrectionResult:
    return next(result for result in report.results if result.rule_id == rule_id)


def test_formula_exact_passed(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="A1", rule_type=RuleType.FORMULA_EXACT, expected_formula="=SUM(1,2)")]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.PASSED


def test_formula_exact_failed(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="A1", rule_type=RuleType.FORMULA_EXACT, expected_formula="=A1+1")]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.FAILED


def test_numeric_value_passed(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="B2", rule_type=RuleType.NUMERIC_VALUE, expected_value=10)]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.PASSED


def test_numeric_value_failed(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="B2", rule_type=RuleType.NUMERIC_VALUE, expected_value=12)]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.FAILED


def test_numeric_value_absolute_tolerance_passed(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    tolerance = ToleranceConfig(mode=ToleranceMode.ABSOLUTE, absolute=0.5)
    profile = build_profile(
        [build_rule(rule_id="r1", cell="B2", rule_type=RuleType.NUMERIC_VALUE, expected_value=10.4, tolerance=tolerance)]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.PASSED


def test_numeric_value_absolute_tolerance_failed(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    tolerance = ToleranceConfig(mode=ToleranceMode.ABSOLUTE, absolute=0.1)
    profile = build_profile(
        [build_rule(rule_id="r1", cell="B2", rule_type=RuleType.NUMERIC_VALUE, expected_value=10.4, tolerance=tolerance)]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.FAILED


def test_text_value_passed(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="C3", rule_type=RuleType.TEXT_VALUE, expected_value="  Hello  ")]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.PASSED


def test_text_value_failed(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="C3", rule_type=RuleType.TEXT_VALUE, expected_value="Hello")]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.FAILED


def test_text_normalized_passed(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="C3", rule_type=RuleType.TEXT_NORMALIZED, expected_value="hello")]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.PASSED


def test_non_empty_passed(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="D4", rule_type=RuleType.NON_EMPTY)]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.PASSED


def test_empty_passed(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="E5", rule_type=RuleType.EMPTY)]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.PASSED


def test_manual_review_status(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="B2", rule_type=RuleType.MANUAL_REVIEW)]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.MANUAL_REVIEW


def test_formula_normalized_not_implemented(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="A1", rule_type=RuleType.FORMULA_NORMALIZED, expected_formula="=SUM(1,2)")]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.WARNING


def test_missing_sheet_produces_error_result(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="A1", rule_type=RuleType.TEXT_VALUE, expected_value="x", sheet_name="MissingSheet")]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.ERROR


def test_disabled_rule_is_skipped(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="B2", rule_type=RuleType.NUMERIC_VALUE, expected_value=10, enabled=False)]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert get_result(report, "r1").status == ResultStatus.SKIPPED


def test_report_contains_coherent_final_grade(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student")
    profile = build_profile(
        [
            build_rule(rule_id="r1", cell="A1", rule_type=RuleType.FORMULA_EXACT, expected_formula="=SUM(1,2)", weight=2.0),
            build_rule(rule_id="r2", cell="B2", rule_type=RuleType.NUMERIC_VALUE, expected_value=11, weight=2.0),
        ],
        max_grade=10.0,
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert report.summary.final_grade == 5.0


def test_supports_xlsx(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student", ".xlsx")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="B2", rule_type=RuleType.NUMERIC_VALUE, expected_value=10)]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert report.student_workbook_format == WorkbookFormat.XLSX


def test_supports_xlsm_without_executing_macros(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student", ".xlsm")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="B2", rule_type=RuleType.NUMERIC_VALUE, expected_value=10)]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert report.student_workbook_format == WorkbookFormat.XLSM
    assert report.macro_enabled is True


def test_no_resource_warning_pattern_and_reader_closes(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student", ".xlsm")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="B2", rule_type=RuleType.NUMERIC_VALUE, expected_value=10)]
    )
    report = CorrectionEngine().correct_workbook(profile, student_path)
    assert report.results[0].status == ResultStatus.PASSED


def test_student_workbook_is_not_modified(tmp_path: Path) -> None:
    student_path = create_student_workbook(tmp_path / "student", ".xlsx")
    profile = build_profile(
        [build_rule(rule_id="r1", cell="A1", rule_type=RuleType.FORMULA_EXACT, expected_formula="=SUM(1,2)")]
    )
    CorrectionEngine().correct_workbook(profile, student_path)
    workbook = load_workbook(student_path, data_only=False)
    try:
        assert workbook["Sheet1"]["A1"].value == "=SUM(1,2)"
    finally:
        workbook.close()
