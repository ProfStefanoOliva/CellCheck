"""Workbook fill color scanner for candidate correction cells."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from cellcheck.models import (
    CellColorMatch,
    ColorScanResult,
    ColorScanSummary,
    WorkbookFormat,
)

from .color_utils import parse_color_input
from .errors import ColorScanError, UnsupportedWorkbookFormatError, WorkbookReadError, WorksheetNotFoundError


class ColorScanner:
    """Scan a workbook for cells whose fill color matches a target RGB color."""

    def __init__(self, workbook_path: str | Path) -> None:
        self.workbook_path = Path(workbook_path)
        self.workbook_format = self._detect_workbook_format(self.workbook_path)
        self.macro_enabled = self.workbook_format == WorkbookFormat.XLSM
        self._workbook = self._open_workbook()

    def __enter__(self) -> "ColorScanner":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    def scan_fill_color(
        self, target_color: str, sheet_names: list[str] | None = None
    ) -> ColorScanResult:
        """Return all cells whose explicit RGB fill matches the target color."""
        color_target = parse_color_input(target_color)
        workbook = self._require_workbook()
        selected_sheet_names = self._resolve_sheet_names(sheet_names)

        matches: list[CellColorMatch] = []
        unsupported_color_cells_count = 0
        ignored_cells_count = 0

        for sheet_name in selected_sheet_names:
            worksheet = workbook[sheet_name]
            max_row = worksheet.max_row or 0
            max_column = worksheet.max_column or 0

            for row in worksheet.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=max_column,
            ):
                for cell in row:
                    color_info = self._extract_rgb_fill(cell)
                    if color_info is None:
                        ignored_cells_count += 1
                        continue

                    color_kind, detected_argb = color_info
                    if color_kind != "rgb":
                        unsupported_color_cells_count += 1
                        continue

                    detected_rgb = detected_argb[-6:]
                    if detected_rgb != color_target.normalized_rgb:
                        ignored_cells_count += 1
                        continue

                    raw_value = cell.value
                    has_formula = isinstance(raw_value, str) and raw_value.startswith("=")
                    matches.append(
                        CellColorMatch(
                            sheet_name=sheet_name,
                            cell=cell.coordinate,
                            detected_rgb=detected_rgb,
                            detected_argb=detected_argb,
                            value_preview=raw_value,
                            has_formula=has_formula,
                            formula=raw_value if has_formula else None,
                        )
                    )

        summary = ColorScanSummary(
            workbook_path=str(self.workbook_path),
            target_rgb=color_target.normalized_rgb,
            target_argb=color_target.normalized_argb,
            scanned_sheets=selected_sheet_names,
            matched_cells_count=len(matches),
            unsupported_color_cells_count=unsupported_color_cells_count,
            ignored_cells_count=ignored_cells_count,
        )
        return ColorScanResult(summary=summary, matches=matches)

    def close(self) -> None:
        """Close workbook and any extra archives used internally by openpyxl."""
        workbook = self._workbook
        if workbook is None:
            return

        extra_archives = []
        for attr_name in ("vba_archive", "_archive"):
            archive = getattr(workbook, attr_name, None)
            if archive is not None and archive not in extra_archives:
                extra_archives.append(archive)

        try:
            workbook.close()
        finally:
            for archive in extra_archives:
                try:
                    archive.close()
                except Exception:
                    pass

            for attr_name in ("vba_archive", "_archive"):
                if hasattr(workbook, attr_name):
                    setattr(workbook, attr_name, None)

            self._workbook = None

    def _open_workbook(self):
        """Open the workbook without read_only so style information stays accessible."""
        try:
            return load_workbook(
                filename=self.workbook_path,
                read_only=False,
                data_only=False,
                keep_vba=self.macro_enabled,
            )
        except Exception as exc:
            raise WorkbookReadError(
                f"Unable to open workbook '{self.workbook_path}' for color scanning."
            ) from exc

    def _resolve_sheet_names(self, sheet_names: list[str] | None) -> list[str]:
        """Validate requested sheets or fall back to all sheets."""
        workbook = self._require_workbook()
        if sheet_names is None:
            return list(workbook.sheetnames)

        missing_sheet_names = [name for name in sheet_names if name not in workbook.sheetnames]
        if missing_sheet_names:
            missing_sheet = missing_sheet_names[0]
            raise WorksheetNotFoundError(
                f"Worksheet '{missing_sheet}' was not found in workbook '{self.workbook_path.name}'."
            )

        return sheet_names

    def _require_workbook(self):
        """Ensure the workbook is still open."""
        if self._workbook is None:
            raise ColorScanError("Workbook is closed.")
        return self._workbook

    @staticmethod
    def _detect_workbook_format(path: Path) -> WorkbookFormat:
        """Validate workbook extension for color scanning."""
        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            return WorkbookFormat.XLSX
        if suffix == ".xlsm":
            return WorkbookFormat.XLSM
        raise UnsupportedWorkbookFormatError(
            f"Unsupported workbook format for '{path}'. Only .xlsx and .xlsm are supported."
        )

    @staticmethod
    def _extract_rgb_fill(cell) -> tuple[str, str] | None:
        """Return explicit fill color information for a cell when available."""
        fill = getattr(cell, "fill", None)
        if fill is None:
            return None

        color = getattr(fill, "fgColor", None)
        if color is None:
            return None

        color_type = getattr(color, "type", None)
        if color_type != "rgb":
            if color_type in {"theme", "indexed", "auto"}:
                return (color_type, "")
            return None

        raw_rgb = getattr(color, "rgb", None)
        if not raw_rgb:
            return None

        normalized = raw_rgb.upper()
        if len(normalized) == 6:
            normalized = f"FF{normalized}"
        if len(normalized) != 8:
            return ("unsupported", "")

        return ("rgb", normalized)
