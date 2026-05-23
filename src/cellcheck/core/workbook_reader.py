"""Prudent workbook reader for .xlsx and .xlsm files."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple

from cellcheck.models import CellSnapshot, WorkbookFormat, WorkbookInfo, WorksheetInfo

from .errors import (
    CellReadError,
    UnsupportedWorkbookFormatError,
    WorkbookReadError,
    WorksheetNotFoundError,
)


class WorkbookReader:
    """Read workbook metadata and cell snapshots without modifying the file."""

    def __init__(self, path: str | Path, data_only: bool = False) -> None:
        self.path = Path(path)
        self.data_only = data_only
        self.workbook_format = self._detect_workbook_format(self.path)
        self.macro_enabled = self.workbook_format == WorkbookFormat.XLSM
        self._workbook = self._open_workbook()

    def __enter__(self) -> "WorkbookReader":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    def get_workbook_info(self) -> WorkbookInfo:
        """Return basic workbook metadata."""
        workbook = self._require_workbook()
        active_sheet = None
        try:
            active_sheet = workbook.active.title
        except Exception:
            active_sheet = None

        return WorkbookInfo(
            path=str(self.path),
            filename=self.path.name,
            workbook_format=self.workbook_format,
            macro_enabled=self.macro_enabled,
            sheet_names=list(workbook.sheetnames),
            active_sheet=active_sheet,
            read_only=True,
            data_only=self.data_only,
        )

    def get_worksheet_info(self, sheet_name: str) -> WorksheetInfo:
        """Return basic worksheet dimensions without scanning the whole sheet."""
        worksheet = self._get_worksheet(sheet_name)

        try:
            calculated_dimension = worksheet.calculate_dimension()
        except Exception:
            calculated_dimension = None

        return WorksheetInfo(
            sheet_name=sheet_name,
            max_row=worksheet.max_row or 0,
            max_column=worksheet.max_column or 0,
            calculated_dimension=calculated_dimension,
        )

    def get_cell_snapshot(self, sheet_name: str, cell_ref: str) -> CellSnapshot:
        """Return a single-cell snapshot from the workbook."""
        worksheet = self._get_worksheet(sheet_name)

        try:
            coordinate_to_tuple(cell_ref)
            cell = worksheet[cell_ref]
        except Exception as exc:
            raise CellReadError(
                f"Invalid or unreadable cell reference '{cell_ref}' in worksheet '{sheet_name}'."
            ) from exc

        raw_value = cell.value
        has_formula = isinstance(raw_value, str) and raw_value.startswith("=")
        formula = raw_value if has_formula and not self.data_only else None

        return CellSnapshot(
            sheet_name=sheet_name,
            cell=cell_ref,
            value=raw_value,
            data_type=getattr(cell, "data_type", None),
            number_format=getattr(cell, "number_format", None),
            has_formula=has_formula,
            formula=formula,
        )

    def close(self) -> None:
        """Close the underlying workbook if it is open."""
        workbook = self._workbook
        if workbook is None:
            return

        # `keep_vba=True` may leave an additional ZipFile-backed archive alive
        # for macro-enabled workbooks. Close both workbook and archive handles
        # explicitly so teardown does not rely on `ZipFile.__del__`.
        extra_archives = []
        for attr_name in ("vba_archive", "_archive"):
            archive = getattr(workbook, attr_name, None)
            if archive is not None and archive not in extra_archives:
                extra_archives.append(archive)

        try:
            workbook.close()
        finally:
            for archive in extra_archives:
                try:
                    archive.close()
                except Exception:
                    pass

            for attr_name in ("vba_archive", "_archive"):
                if hasattr(workbook, attr_name):
                    setattr(workbook, attr_name, None)

            self._workbook = None

    def _open_workbook(self):
        """Open the workbook in read-only mode."""
        try:
            return load_workbook(
                filename=self.path,
                read_only=True,
                data_only=self.data_only,
                keep_vba=self.macro_enabled,
            )
        except Exception as exc:
            raise WorkbookReadError(
                f"Unable to open workbook '{self.path}' in read-only mode."
            ) from exc

    def _get_worksheet(self, sheet_name: str):
        """Return a worksheet or raise a domain-specific error."""
        workbook = self._require_workbook()
        if sheet_name not in workbook.sheetnames:
            raise WorksheetNotFoundError(
                f"Worksheet '{sheet_name}' was not found in workbook '{self.path.name}'."
            )
        return workbook[sheet_name]

    def _require_workbook(self):
        """Ensure the workbook is still open."""
        if self._workbook is None:
            raise WorkbookReadError("Workbook is closed.")
        return self._workbook

    @staticmethod
    def _detect_workbook_format(path: Path) -> WorkbookFormat:
        """Map a file extension to the supported workbook format enum."""
        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            return WorkbookFormat.XLSX
        if suffix == ".xlsm":
            return WorkbookFormat.XLSM
        raise UnsupportedWorkbookFormatError(
            f"Unsupported workbook format for '{path}'. Only .xlsx and .xlsm are supported."
        )
