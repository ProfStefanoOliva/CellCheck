"""Synthetic workbook generation helpers for manual CellCheck testing."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUTPUT_DIR = Path(__file__).resolve().parents[3] / "manual_tests" / "generated"
TARGET_COLOR_RGB = "D9D9D9"
TARGET_COLOR_ARGB = f"FF{TARGET_COLOR_RGB}"
SHEET_NAME = "Esercizio"
TABLE_START_ROW = 4

HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF1F6AA5")
TARGET_FILL = PatternFill(fill_type="solid", fgColor=TARGET_COLOR_ARGB)
INFO_FILL = PatternFill(fill_type="solid", fgColor="FF223044")
THIN_BORDER = Border(
    left=Side(style="thin", color="FF2F3D4C"),
    right=Side(style="thin", color="FF2F3D4C"),
    top=Side(style="thin", color="FF2F3D4C"),
    bottom=Side(style="thin", color="FF2F3D4C"),
)
HEADER_FONT = Font(color="FFFFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True)
ITALIC_FONT = Font(italic=True, color="FF44515E")


@dataclass(frozen=True)
class ExerciseCase:
    """Declarative description of one synthetic exercise row."""

    code: str
    label: str
    description: str
    category: str
    input_1: Any = None
    input_2: Any = None


@dataclass(frozen=True)
class WorkbookSpec:
    """Workbook generation configuration."""

    filename: str
    title: str
    response_map: dict[str, Any]
    cases: list[ExerciseCase] = field(default_factory=list)
    is_macro_container: bool = False


AUTOMATIC_CASES: list[ExerciseCase] = [
    ExerciseCase(
        code="FORMULA_OK",
        label="Controllo automatico - formula somma",
        description="Percorso automatico: inserire la formula di somma corretta.",
        category="controllo automatico",
        input_1=10,
        input_2=15,
    ),
    ExerciseCase(
        code="FORMULA_WRONG",
        label="Controllo automatico - formula prodotto",
        description="Percorso automatico: inserire la formula di prodotto corretta.",
        category="controllo automatico",
        input_1=6,
        input_2=7,
    ),
    ExerciseCase(
        code="NUMERIC_OK",
        label="Controllo automatico - valore esatto",
        description="Percorso automatico: inserire il numero esatto richiesto.",
        category="controllo automatico",
        input_1="Target",
        input_2="",
    ),
    ExerciseCase(
        code="NUMERIC_TOLERANCE",
        label="Controllo automatico - tolleranza",
        description="Percorso automatico: numero usato per verifiche con tolleranza.",
        category="controllo automatico",
        input_1="Atteso",
        input_2="100 +/- tolleranza",
    ),
    ExerciseCase(
        code="TEXT_OK",
        label="Controllo automatico - testo esatto",
        description="Percorso automatico: inserire il testo esatto atteso.",
        category="controllo automatico",
        input_1="Etichetta",
        input_2="",
    ),
    ExerciseCase(
        code="TEXT_WRONG",
        label="Controllo automatico - testo rigido",
        description="Percorso automatico: inserire il testo esatto per controllo rigido.",
        category="controllo automatico",
        input_1="Conferma",
        input_2="",
    ),
    ExerciseCase(
        code="TEXT_NORMALIZED",
        label="Controllo automatico - testo normalizzato",
        description="Percorso automatico: caso utile per controlli text_normalized.",
        category="controllo automatico",
        input_1="Normalizzare",
        input_2="",
    ),
    ExerciseCase(
        code="NON_EMPTY_EXPECTED",
        label="Controllo automatico - cella non vuota attesa",
        description="Percorso automatico: la cella deve contenere un valore.",
        category="controllo automatico",
        input_1="Compilare",
        input_2="",
    ),
]

MANUAL_REVIEW_CASES: list[ExerciseCase] = [
    ExerciseCase(
        code="EMPTY_EXPECTED",
        label="Revisione manuale - cella vuota attesa",
        description=(
            "Caso didattico separato: la soluzione resta vuota e il profilo generato richiede revisione manuale del docente."
        ),
        category="revisione manuale",
        input_1="Lasciare vuoto",
        input_2="Controllo docente richiesto",
    ),
    ExerciseCase(
        code="MANUAL_REVIEW",
        label="Revisione manuale - controllo docente",
        description=(
            "Caso dedicato alla revisione manuale: la soluzione resta volutamente vuota per generare una regola manual_review."
        ),
        category="revisione manuale",
        input_1="Osservazione",
        input_2="Controllo docente richiesto",
    ),
]

ALL_CASES = AUTOMATIC_CASES + MANUAL_REVIEW_CASES


SOLUTION_VALUES: dict[str, Any] = {
    "FORMULA_OK": lambda row: f"=C{row}+D{row}",
    "FORMULA_WRONG": lambda row: f"=C{row}*D{row}",
    "NUMERIC_OK": 42,
    "NUMERIC_TOLERANCE": 100,
    "TEXT_OK": "Bilancio",
    "TEXT_WRONG": "Confermato",
    "TEXT_NORMALIZED": "Risposta Normalizzata",
    "EMPTY_EXPECTED": None,
    "NON_EMPTY_EXPECTED": "Presente",
    "MANUAL_REVIEW": None,
}

AUTOMATIC_SOLUTION_VALUES = {
    code: value for code, value in SOLUTION_VALUES.items() if code != "MANUAL_REVIEW"
}

WORKBOOK_SPECS: list[WorkbookSpec] = [
    WorkbookSpec(
        filename="01_modello_vuoto.xlsx",
        title="Modello vuoto per profilo automatico",
        response_map={},
        cases=AUTOMATIC_CASES,
    ),
    WorkbookSpec(
        filename="02_modello_risolto.xlsx",
        title="Modello risolto per profilo automatico",
        response_map=AUTOMATIC_SOLUTION_VALUES,
        cases=AUTOMATIC_CASES,
    ),
    WorkbookSpec(
        filename="03_studente_perfetto_automatico.xlsx",
        title="Studente perfetto per controlli automatici",
        response_map=AUTOMATIC_SOLUTION_VALUES,
        cases=AUTOMATIC_CASES,
    ),
    WorkbookSpec(
        filename="04_studente_errori_misti.xlsx",
        title="Studente con errori misti",
        response_map={
            "FORMULA_OK": lambda row: f"=C{row}+D{row}",
            "FORMULA_WRONG": lambda row: f"=C{row}+D{row}",
            "NUMERIC_OK": 42,
            "NUMERIC_TOLERANCE": 120,
            "TEXT_OK": "Bilancio",
            "TEXT_WRONG": "Respinto",
            "TEXT_NORMALIZED": "  RISPOSTA NORMALIZZATA  ",
            "EMPTY_EXPECTED": "Non doveva esserci",
            "NON_EMPTY_EXPECTED": None,
        },
        cases=AUTOMATIC_CASES,
    ),
    WorkbookSpec(
        filename="05_studente_parziale.xlsx",
        title="Studente con risposte parziali",
        response_map={
            "FORMULA_OK": lambda row: f"=C{row}+D{row}",
            "FORMULA_WRONG": None,
            "NUMERIC_OK": 42,
            "NUMERIC_TOLERANCE": None,
            "TEXT_OK": "Bilancio",
            "TEXT_WRONG": None,
            "TEXT_NORMALIZED": " risposta normalizzata ",
            "EMPTY_EXPECTED": None,
            "NON_EMPTY_EXPECTED": "Abbozzo",
        },
        cases=AUTOMATIC_CASES,
    ),
    WorkbookSpec(
        filename="06_modello_vuoto_revisione_manuale.xlsx",
        title="Modello vuoto dedicato alla revisione manuale",
        response_map={},
        cases=ALL_CASES,
    ),
    WorkbookSpec(
        filename="07_modello_risolto_revisione_manuale.xlsx",
        title="Modello risolto dedicato alla revisione manuale",
        response_map=SOLUTION_VALUES,
        cases=ALL_CASES,
    ),
    WorkbookSpec(
        filename="08_studente_revisione_manuale.xlsx",
        title="Studente dedicato alla revisione manuale",
        response_map={
            **AUTOMATIC_SOLUTION_VALUES,
            "MANUAL_REVIEW": "Da verificare a vista",
        },
        cases=ALL_CASES,
    ),
    WorkbookSpec(
        filename="09_modello_vuoto_macro.xlsm",
        title="Modello vuoto macro-enabled strutturale",
        response_map={},
        cases=AUTOMATIC_CASES,
        is_macro_container=True,
    ),
    WorkbookSpec(
        filename="10_studente_macro.xlsm",
        title="Studente macro-enabled strutturale",
        response_map={
            "FORMULA_OK": lambda row: f"=C{row}+D{row}",
            "FORMULA_WRONG": lambda row: f"=C{row}+D{row}",
            "NUMERIC_OK": 42,
            "NUMERIC_TOLERANCE": 103,
            "TEXT_OK": "Bilancio",
            "TEXT_WRONG": "Confermato",
            "TEXT_NORMALIZED": " risposta normalizzata ",
            "EMPTY_EXPECTED": None,
            "NON_EMPTY_EXPECTED": "Presente",
        },
        cases=AUTOMATIC_CASES,
        is_macro_container=True,
    ),
]


def get_output_dir() -> Path:
    """Return the default output directory for synthetic workbooks."""
    return OUTPUT_DIR


def generate_all_workbooks(output_dir: Path | None = None) -> list[Path]:
    """Generate every synthetic workbook and return the written paths."""
    target_dir = output_dir or OUTPUT_DIR
    target_dir.mkdir(parents=True, exist_ok=True)
    return [generate_workbook(spec, target_dir) for spec in WORKBOOK_SPECS]


def generate_workbook(spec: WorkbookSpec, output_dir: Path) -> Path:
    """Build and save one workbook according to the provided specification."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME

    configure_sheet_layout(sheet, spec)
    populate_sheet_rows(sheet, spec.cases, spec.response_map)

    output_path = output_dir / spec.filename
    workbook.save(output_path)
    workbook.close()
    return output_path


def configure_sheet_layout(sheet, spec: WorkbookSpec) -> None:
    """Create a consistent workbook layout for manual tests."""
    sheet["A1"] = "CellCheck - Workbook sintetico per test manuali"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = spec.title
    sheet["A2"].font = Font(bold=True)
    sheet["A3"] = (
        "Le celle colorate in colonna F sono candidate alla correzione automatica per il profilo collegato a questo workbook."
    )
    sheet["A3"].font = ITALIC_FONT

    headers = ["Caso", "Categoria", "Descrizione", "Dato 1", "Dato 2", "Risposta"]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=TABLE_START_ROW, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = f"A{TABLE_START_ROW + 1}"
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["C"].width = 48
    sheet.column_dimensions["D"].width = 16
    sheet.column_dimensions["E"].width = 22
    sheet.column_dimensions["F"].width = 30
    sheet.column_dimensions["H"].width = 24
    sheet.column_dimensions["I"].width = 56

    profile_mode = (
        "profilo dedicato alla revisione manuale"
        if spec.cases == ALL_CASES
        else "profilo per controlli automatici puri"
    )
    info_pairs = [
        ("Formato file", spec.filename.split(".")[-1].lower()),
        ("Colore target", f"#{TARGET_COLOR_RGB}"),
        ("Foglio principale", SHEET_NAME),
        ("Scenario didattico", spec.title),
        ("Profilo atteso", profile_mode),
    ]
    if spec.is_macro_container:
        info_pairs.append(
            (
                "Nota .xlsm",
                "Contenitore strutturale per test prudente del formato macro-enabled; nessuna macro VBA reale inclusa.",
            )
        )

    info_row = 1
    for label, value in info_pairs:
        label_cell = sheet.cell(row=info_row, column=8, value=label)
        value_cell = sheet.cell(row=info_row, column=9, value=value)
        label_cell.fill = INFO_FILL
        label_cell.font = HEADER_FONT
        label_cell.border = THIN_BORDER
        value_cell.border = THIN_BORDER
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        info_row += 1


def populate_sheet_rows(
    sheet,
    cases: list[ExerciseCase],
    response_map: dict[str, Any],
) -> None:
    """Populate all exercise cases and apply the target fill to answer cells."""
    first_data_row = TABLE_START_ROW + 1
    for offset, case in enumerate(cases):
        row = first_data_row + offset
        answer_value = resolve_case_value(response_map, case.code, row)

        row_values = [
            case.label,
            case.category,
            case.description,
            case.input_1,
            case.input_2,
            answer_value,
        ]

        for column_index, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row, column=column_index, value=value)
            cell.border = THIN_BORDER
            if column_index == 6:
                cell.fill = TARGET_FILL
            if column_index in (1, 2, 3):
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def resolve_case_value(response_map: dict[str, Any], case_code: str, row: int) -> Any:
    """Resolve the value to place in a response cell for one workbook variant."""
    value = response_map.get(case_code)
    if callable(value):
        return value(row)
    return value
